import openpyxl

def leer_excel(nombre_archivo):
    workbook = openpyxl.load_workbook(nombre_archivo)
    sheet = workbook["1"]
    valor = sheet.cell(8,3).value
    cuentas = []

    for fila in range(8,36):
        variable_nombre = sheet.cell(fila,2).value
        cuenta = {
            "nombre": variable_nombre,
            "totales":[],
            "variaciones":[]
        }
        for columna in range(3, 15):
            variable_valor = sheet.cell(fila, columna).value
            cuenta["totales"].append(variable_valor) 

        cuentas.append(cuenta)
    return cuentas
def calcular_variacion(valor1, valor2):
    variacion = ((valor2 - valor1) / valor1) * 100
    return variacion
def crear_excel(nombre_archivo, cuentas):
    workbook = openpyxl.Workbook()
    sheet = workbook.create_sheet("resultado")
    fila = 1
    for cuenta in cuentas:
        columna = 1
        sheet.cell(fila, columna,cuenta["nombre"])
        for variacion in cuenta["variaciones"]:
            columna += 1
            sheet.cell(fila, columna,variacion)
        fila += 1
    workbook.save(nombre_archivo)

cuentas = leer_excel("clase14/MES 5 - Noviembre 2025 - Datos Financieros.xlsx")
for cuenta in cuentas:
    for indice in range(0, len(cuenta["totales"])-2):
        variacion = calcular_variacion(cuenta["totales"][indice],cuenta["totales"][indice + 1])
        cuenta["variaciones"].append(variacion)
print(cuentas)
crear_excel("clase14/resultado.xlsx", cuentas)