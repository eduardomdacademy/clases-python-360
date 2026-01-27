import openpyxl

def calcular_ultima_columna(sheet):
    columna_inicial = 3
    columna_actual = columna_inicial
    fila_encabezado=6
    while(True):
        valorEncabezado = sheet.cell(fila_encabezado, columna_actual).value
        if valorEncabezado == "Separador":
            return columna_actual+1
        columna_actual += 1

def leer_excel(nombre_archivo):
    workbook = openpyxl.load_workbook(nombre_archivo)
    sheet = workbook["1"]
    cuentas = []

    for fila in range(8,sheet.max_row):
        variable_nombre = sheet.cell(fila,2).value
        if variable_nombre == "" or variable_nombre is None:
            continue
        cuenta = {
            "nombre": variable_nombre,
            "totales":[],
            "variaciones":[]
        }

        for columna in range(3, calcular_ultima_columna(sheet)):
            variable_valor = sheet.cell(fila, columna).value
            cuenta["totales"].append(variable_valor) 

        cuentas.append(cuenta)
    return cuentas
def calcular_variacion(valor1, valor2):
    if valor1 == 0 or valor1 is None or valor2 is None:
        return ""
    variacion = ((valor2 - valor1) / valor1) * 100
    return variacion
def crear_excel(nombre_archivo, cuentas):
    workbook = openpyxl.Workbook()
    sheet = workbook.create_sheet("resultado")
    fila = 1
    for cuenta in cuentas:
        columna = 1
        sheet.cell(fila, columna,cuenta["nombre"])
        for variacion in cuenta["variaciones"]:
            columna += 1
            sheet.cell(fila, columna,variacion)
        fila += 1
    workbook.save(nombre_archivo)

cuentas = leer_excel("clase15/continuacion_excel_bcp/bcp_datos_financieros_noviembre_2025.xlsx")
for cuenta in cuentas:
    for indice in range(0, len(cuenta["totales"])-2):
        variacion = calcular_variacion(cuenta["totales"][indice],cuenta["totales"][indice + 1])
        cuenta["variaciones"].append(variacion)
print(cuentas)
crear_excel("clase15/continuacion_excel_bcp/resultado.xlsx", cuentas)