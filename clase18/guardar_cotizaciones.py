import openpyxl
from cotizacion import Cotizacion

def guardar_resultados(cotizaciones, path_archivo_resultado):
    cotizaciones.sort()
    
    workbook = openpyxl.Workbook()
    monedas_set = {cotizaciones[0].moneda}
    for cotizacion in cotizaciones:
        cotizacion:Cotizacion
        print(cotizacion.fecha)
        monedas_set.add(cotizacion.moneda)
    print(monedas_set)
    for moneda in monedas_set:
        workbook.create_sheet(moneda)
        sheet = workbook[moneda]
        sheet.cell(1,1,"Fecha")
        sheet.cell(1,2,"Compra")
        sheet.cell(1,3,"Venta")        
        fila = 2
        for cotizacion in cotizaciones:
            cotizacion:Cotizacion
            if cotizacion.moneda != moneda:
                continue
            sheet.cell(fila,1,cotizacion.fecha)
            sheet.cell(fila,2,cotizacion.compra)
            sheet.cell(fila,3,cotizacion.venta)
            fila += 1
    del workbook["Sheet"]
    workbook.save(path_archivo_resultado)
