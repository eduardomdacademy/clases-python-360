from pathlib import Path
import openpyxl
from cotizacion import Cotizacion
def leer_excel(path: Path):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook["Cotizaciones Diarias"]
    
    encabezado_mes_año = sheet.cell(8,2).value
    mes_año = encabezado_mes_año.split("/")
    mes_letra = mes_año[0]
    año = mes_año[1]
    mes_numero = mes_letra_a_numero(mes_letra)
    print(año,"-",mes_numero)
    fila_moneda = 0

    resultado = []
    for fila in range(1,sheet.max_row):
        dia_letra = str(sheet.cell(fila, 1).value)
        if dia_letra is None or dia_letra.isnumeric() == False:
            continue
        if dia_letra == "1":
            fila_moneda = fila -2
        fecha = año + "-" + mes_numero + "-" +dia_letra
        for columna in range(2, 12,2):
            moneda = sheet.cell(fila_moneda, columna).value
            compra = sheet.cell(fila, columna).value
            venta = sheet.cell(fila, columna + 1).value
            if compra == "-" or compra == 0:
                continue
            if venta == "-" or venta == 0:
                continue
            cotizacion = Cotizacion(fecha, moneda, compra, venta)
            resultado.append(cotizacion)
            # cotizacion.imprimir()
    return resultado



def mes_letra_a_numero(mes:str):
    mesMinuscula = mes.lower()
    if mesMinuscula == "enero": return "01"
    if mesMinuscula == "febrero": return "02"
    if mesMinuscula == "marzo": return "03"
    if mesMinuscula == "abril": return "04"
    if mesMinuscula == "mayo": return "05"
    if mesMinuscula == "junio": return "06"
    if mesMinuscula == "julio": return "07"
    if mesMinuscula == "agosto": return "08"
    if mesMinuscula == "septiembre": return "09"
    if mesMinuscula == "setiembre": return "09"
    if mesMinuscula == "octubre": return "10"
    if mesMinuscula == "noviembre": return "11"
    if mesMinuscula == "diciembre": return "12"
    
    