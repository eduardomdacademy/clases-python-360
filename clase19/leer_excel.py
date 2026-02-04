import openpyxl
from Cotizacion import Cotizacion
def leer_excel(path_archivo):
    workbook = openpyxl.open(path_archivo)
    sheet = workbook["Dólar"]
    resultado = []
    for fila in range(2,sheet.max_row):
        fecha = sheet.cell(fila, 1).value
        compra = sheet.cell(fila,2).value
        venta = sheet.cell(fila, 3).value
        cotizacion = Cotizacion(fecha, compra, venta)
        resultado.append(cotizacion)
    return resultado