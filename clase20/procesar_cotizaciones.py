import openpyxl
from openpyxl.drawing.image import Image as XlImage
from pathlib import Path
from datetime import date, datetime, timedelta
from io import BytesIO
import matplotlib.pyplot as plt
import matplotlib.dates as mdates

MESES = {
    'enero': 1, 'febrero': 2, 'marzo': 3, 'abril': 4,
    'mayo': 5, 'junio': 6, 'julio': 7, 'agosto': 8,
    'septiembre': 9, 'octubre': 10, 'noviembre': 11, 'diciembre': 12,
}


def parsear_mes_anio(mes_anio):
    """Convierte 'Agosto/2025' en (8, 2025)."""
    partes = mes_anio.replace("/", " ").split()
    nombre_mes = partes[0].strip().lower()
    anio = int(partes[1].strip())
    return MESES[nombre_mes], anio


def es_valido(valor):
    """Retorna True si el valor no es None, vacío ni '-'."""
    return valor is not None and valor != '' and valor != '-'


def extraer_monedas(ws, fila_monedas, fila_tipo, max_col):
    """
    Extrae las monedas y sus columnas de compra/venta de una tabla.

    Args:
        ws: Hoja de trabajo de openpyxl
        fila_monedas: Número de fila que contiene los nombres de monedas (1-indexed)
        fila_tipo: Número de fila que contiene "Compra"/"Venta" (1-indexed)
        max_col: Número máximo de columnas a recorrer

    Returns:
        Diccionario con {nombre_moneda: {'compra': col_idx, 'venta': col_idx}}
    """
    monedas = {}
    for col_idx in range(2, max_col + 1):
        moneda = ws.cell(row=fila_monedas, column=col_idx).value
        tipo = ws.cell(row=fila_tipo, column=col_idx).value

        if es_valido(moneda):
            moneda_nombre = str(moneda).strip()
            if moneda_nombre not in monedas:
                monedas[moneda_nombre] = {'compra': None, 'venta': None}

            if tipo is not None:
                tipo_str = str(tipo).strip().lower()
                if 'compra' in tipo_str:
                    monedas[moneda_nombre]['compra'] = col_idx
                elif 'venta' in tipo_str:
                    monedas[moneda_nombre]['venta'] = col_idx
        elif tipo is not None:
            tipo_str = str(tipo).strip().lower()
            if monedas and 'venta' in tipo_str:
                ultima_moneda = list(monedas.keys())[-1]
                monedas[ultima_moneda]['venta'] = col_idx

    return monedas


def procesar_datos_moneda(ws, fila_datos_inicio, cols, mes, anio):
    """
    Procesa los datos de cotización de una moneda.

    Args:
        ws: Hoja de trabajo de openpyxl
        fila_datos_inicio: Número de fila de inicio de datos (1-indexed)
        cols: Diccionario con {'compra': col_idx, 'venta': col_idx}
        mes: Número de mes (1-12)
        anio: Año (ej. 2025)

    Returns:
        Lista de diccionarios con {Fecha, Compra, Venta}
    """
    datos_moneda = []
    fila_idx = fila_datos_inicio

    while True:
        dia = ws.cell(row=fila_idx, column=1).value

        if dia is None:
            break

        try:
            dia_num = int(dia)
        except (ValueError, TypeError):
            fila_idx += 1
            continue

        compra = ws.cell(row=fila_idx, column=cols['compra']).value if cols['compra'] else None
        venta = ws.cell(row=fila_idx, column=cols['venta']).value if cols['venta'] else None

        if compra is not None or venta is not None:
            if compra == '-':
                compra = None
            if venta == '-':
                venta = None

            datos_moneda.append({
                'Fecha': date(anio, mes, dia_num),
                'Compra': compra,
                'Venta': venta,
            })

        fila_idx += 1

    return datos_moneda


def procesar_tabla(ws, fila_monedas, fila_tipo, fila_datos_inicio, mes, anio):
    """
    Procesa una tabla completa de cotizaciones.

    Args:
        ws: Hoja de trabajo de openpyxl
        fila_monedas: Fila con nombres de monedas (1-indexed)
        fila_tipo: Fila con "Compra"/"Venta" (1-indexed)
        fila_datos_inicio: Primera fila con datos (1-indexed)
        mes: Número de mes (1-12)
        anio: Año (ej. 2025)

    Returns:
        Diccionario con {nombre_moneda: [datos]}
    """
    monedas = extraer_monedas(ws, fila_monedas, fila_tipo, ws.max_column)

    resultado = {}
    for moneda, cols in monedas.items():
        datos = procesar_datos_moneda(ws, fila_datos_inicio, cols, mes, anio)
        if datos:
            resultado[moneda] = datos

    return resultado


def procesar_archivo(archivo):
    """
    Procesa un archivo Excel de cotizaciones y retorna un diccionario
    con {nombre_moneda: [datos]} para todas las monedas encontradas.
    """
    print(f"\nLeyendo archivo: {archivo.name}")

    wb = openpyxl.load_workbook(archivo, data_only=True)
    ws = wb["Cotizaciones Diarias"]
    print(f"  Filas: {ws.max_row}, Columnas: {ws.max_column}")

    # Buscar el mes/año en las primeras filas
    mes_anio = None
    for fila in range(1, min(11, ws.max_row + 1)):
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=fila, column=col).value
            if val is not None and isinstance(val, str) and "/" in val:
                mes_anio = val
                print(f"  Mes/Año encontrado: {mes_anio}")
                break
        if mes_anio:
            break

    if not mes_anio:
        print(f"  Advertencia: No se encontró mes/año en {archivo.name}, saltando.")
        wb.close()
        return {}

    # Configuración de las tres tablas (filas 1-indexed para openpyxl)
    tablas = [
        {'fila_monedas': 10, 'fila_tipo': 11, 'fila_datos_inicio': 12, 'nombre': 'Primera tabla'},
        {'fila_monedas': 60, 'fila_tipo': 61, 'fila_datos_inicio': 62, 'nombre': 'Segunda tabla'},
        {'fila_monedas': 109, 'fila_tipo': 110, 'fila_datos_inicio': 111, 'nombre': 'Tercera tabla'},
    ]

    mes, anio = parsear_mes_anio(mes_anio)

    monedas_archivo = {}
    for config in tablas:
        print(f"  Procesando {config['nombre']}...")
        monedas_tabla = procesar_tabla(
            ws,
            config['fila_monedas'],
            config['fila_tipo'],
            config['fila_datos_inicio'],
            mes,
            anio,
        )

        for moneda, datos in monedas_tabla.items():
            if moneda in monedas_archivo:
                monedas_archivo[moneda].extend(datos)
            else:
                monedas_archivo[moneda] = datos

    wb.close()
    return monedas_archivo


def suavizado_holt(valores, alfa=0.3, beta=0.1, dias_proyeccion=22):
    """
    Suavizado Exponencial Doble (Holt) para proyectar valores futuros.

    Args:
        valores: Lista de valores numéricos históricos.
        alfa: Factor de suavizado del nivel (0-1).
        beta: Factor de suavizado de la tendencia (0-1).
        dias_proyeccion: Cantidad de días a proyectar.

    Returns:
        Lista con los valores proyectados.
    """
    if len(valores) < 2:
        return []

    # Inicializar nivel y tendencia
    nivel = valores[0]
    tendencia = valores[1] - valores[0]

    # Recorrer datos históricos para ajustar nivel y tendencia
    for valor in valores[1:]:
        nivel_anterior = nivel
        nivel = alfa * valor + (1 - alfa) * (nivel_anterior + tendencia)
        tendencia = beta * (nivel - nivel_anterior) + (1 - beta) * tendencia

    # Generar proyección
    proyeccion = []
    for i in range(1, dias_proyeccion + 1):
        proyeccion.append(nivel + i * tendencia)

    return proyeccion


def generar_fechas_futuras(ultima_fecha, cantidad):
    """
    Genera fechas futuras hábiles (lunes a viernes) a partir de una fecha.

    Args:
        ultima_fecha: Última fecha de los datos reales.
        cantidad: Cantidad de días hábiles a generar.

    Returns:
        Lista de objetos date.
    """
    fechas = []
    fecha_actual = ultima_fecha
    while len(fechas) < cantidad:
        fecha_actual += timedelta(days=1)
        # 0=lunes, 5=sábado, 6=domingo
        if fecha_actual.weekday() < 5:
            fechas.append(fecha_actual)
    return fechas


def generar_grafico(datos_moneda, nombre_moneda):
    """
    Genera un gráfico de línea con las cotizaciones de compra y venta.

    Returns:
        BytesIO con la imagen PNG del gráfico.
    """
    # Filtrar puntos con datos válidos para que la línea sea continua
    fechas_compra = [r['Fecha'] for r in datos_moneda if r['Compra'] is not None]
    compras = [r['Compra'] for r in datos_moneda if r['Compra'] is not None]
    fechas_venta = [r['Fecha'] for r in datos_moneda if r['Venta'] is not None]
    ventas = [r['Venta'] for r in datos_moneda if r['Venta'] is not None]

    fig, ax = plt.subplots(figsize=(14, 5))

    ax.plot(fechas_compra, compras, label='Compra', color='#2196F3', linewidth=1.5)
    ax.plot(fechas_venta, ventas, label='Venta', color='#F44336', linewidth=1.5)

    # Proyección con Suavizado Exponencial Doble (Holt)
    dias_proyeccion = 22
    ultima_fecha = max(fechas_compra[-1:] + fechas_venta[-1:])
    fechas_futuras = generar_fechas_futuras(ultima_fecha, dias_proyeccion)

    if compras:
        proy_compra = suavizado_holt(compras, dias_proyeccion=dias_proyeccion)
        ax.plot(
            [fechas_compra[-1]] + fechas_futuras,
            [compras[-1]] + proy_compra,
            label='Compra (estimado)', color='#2196F3', linewidth=1.5, linestyle='--',
        )

    if ventas:
        proy_venta = suavizado_holt(ventas, dias_proyeccion=dias_proyeccion)
        ax.plot(
            [fechas_venta[-1]] + fechas_futuras,
            [ventas[-1]] + proy_venta,
            label='Venta (estimado)', color='#F44336', linewidth=1.5, linestyle='--',
        )

    # Línea vertical separando datos reales de la proyección
    ax.axvline(x=ultima_fecha, color='gray', linestyle=':', linewidth=1, alpha=0.7)

    ax.set_title(f'Cotización {nombre_moneda} - Compra / Venta', fontsize=14)
    ax.set_xlabel('Fecha')
    ax.set_ylabel('Cotización (Gs.)')
    ax.legend()
    ax.grid(True, alpha=0.3)

    ax.xaxis.set_major_locator(mdates.MonthLocator())
    ax.xaxis.set_major_formatter(mdates.DateFormatter('%b %Y'))
    fig.autofmt_xdate(rotation=45)

    fig.tight_layout()

    buf = BytesIO()
    fig.savefig(buf, format='png', dpi=120)
    plt.close(fig)
    buf.seek(0)
    return buf


def procesar_cotizaciones():
    """
    Lee todos los archivos Excel de cotizaciones en la carpeta y genera un nuevo Excel
    con una hoja por cada moneda, mostrando la cotización de compra/venta por día.
    """

    carpeta_cotizacion = Path("clase20") / "cotizacion"

    # Buscar todos los archivos Excel que no sean temporales (~$)
    archivos = sorted([
        f for f in carpeta_cotizacion.glob("Cotizaciones - *.xlsx")
        if not f.name.startswith("~$")
    ])

    if not archivos:
        print(f"Error: No se encontraron archivos de cotizaciones en {carpeta_cotizacion}")
        return

    print(f"Se encontraron {len(archivos)} archivo(s) de cotizaciones:")
    for a in archivos:
        print(f"  - {a.name}")

    try:
        todas_monedas = {}

        for archivo in archivos:
            monedas_archivo = procesar_archivo(archivo)

            for moneda, datos in monedas_archivo.items():
                if moneda in todas_monedas:
                    todas_monedas[moneda].extend(datos)
                else:
                    todas_monedas[moneda] = datos

        # Ordenar los datos de cada moneda por fecha
        for moneda in todas_monedas:
            todas_monedas[moneda].sort(key=lambda r: r['Fecha'])

        print(f"\nTotal de monedas encontradas: {len(todas_monedas)}")
        print(f"Monedas: {list(todas_monedas.keys())}")

        # Crear archivo Excel de salida
        carpeta_resultados = carpeta_cotizacion.parent / "resultados"
        carpeta_resultados.mkdir(exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        archivo_salida = carpeta_resultados / f"Cotizaciones_Por_Moneda_Todos_{timestamp}.xlsx"

        print(f"\nGenerando archivo Excel de salida...")
        wb_salida = openpyxl.Workbook()
        # Eliminar la hoja por defecto
        wb_salida.remove(wb_salida.active)

        for moneda, datos_moneda in todas_monedas.items():
            nombre_hoja = moneda[:31]
            ws_salida = wb_salida.create_sheet(title=nombre_hoja)

            # Escribir encabezados
            ws_salida.cell(row=1, column=1, value='Fecha')
            ws_salida.cell(row=1, column=2, value='Compra')
            ws_salida.cell(row=1, column=3, value='Venta')

            # Escribir datos
            for i, registro in enumerate(datos_moneda, start=2):
                ws_salida.cell(row=i, column=1, value=registro['Fecha'])
                ws_salida.cell(row=i, column=2, value=registro['Compra'])
                ws_salida.cell(row=i, column=3, value=registro['Venta'])

            # Insertar gráfico de línea
            buf = generar_grafico(datos_moneda, moneda)
            img = XlImage(buf)
            ws_salida.add_image(img, 'E2')

            print(f"  - {moneda}: {len(datos_moneda)} registros guardados")

        wb_salida.save(archivo_salida)
        wb_salida.close()

        print(f"\nArchivo generado exitosamente: {archivo_salida}")

    except Exception as e:
        print(f"Error al procesar el archivo: {e}")
        import traceback
        traceback.print_exc()
        return


if __name__ == "__main__":
    procesar_cotizaciones()
