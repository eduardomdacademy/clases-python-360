from openpyxl import load_workbook, Workbook
from datetime import datetime
from pathlib import Path
MESES = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
         "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]

MESES_ABREV = {"Ene": "Enero", "Feb": "Febrero", "Mar": "Marzo", "Abr": "Abril",
               "May": "Mayo", "Jun": "Junio", "Jul": "Julio", "Ago": "Agosto",
               "Set": "Septiembre", "Oct": "Octubre", "Nov": "Noviembre", "Dic": "Diciembre"}

# Leer el archivo fuente
path = Path("clase21") / "cotizacion" / "Cotizaciones - Agosto 2025.xlsx"
wb = load_workbook(path, data_only=True)
ws = wb.worksheets[1]  # Segunda hoja: "USD Fin Mes"

# Extraer datos de la primera tabla (anual) que empieza en fila 7, columnas 1-3
datos = []
for fila in range(7, ws.max_row + 1):
    valores = []
    for col in range(1, 4):
        valores.append(ws.cell(row=fila, column=col).value)
    if all(v is None for v in valores):
        break
    datos.append(valores)

# Extraer datos de las tablas en fila 16 (años 1970-1974)
# Años en columnas 5, 8, 11, 14, 17 - meses abreviados en columna 4
COL_ANIOS_F16 = [5, 8, 11, 14, 17]
anios_f16 = {}
for col_anio in COL_ANIOS_F16:
    valor = ws.cell(row=16, column=col_anio).value
    if isinstance(valor, (int, float)):
        anios_f16[col_anio] = int(valor)

for fila_mes in range(19, 31):
    mes_abrev = ws.cell(row=fila_mes, column=4).value
    if mes_abrev not in MESES_ABREV:
        continue
    mes = MESES_ABREV[mes_abrev]
    for col_anio, anio in anios_f16.items():
        compra = ws.cell(row=fila_mes, column=col_anio).value
        venta = ws.cell(row=fila_mes, column=col_anio + 1).value
        if compra is not None:
            datos.append([f"{anio} {mes}", compra, venta])

# Extraer datos de las tablas mensuales (fila 33 en adelante)
# Hay bloques de 6 tablas lado a lado, con años en columnas 2, 5, 8, 11, 14, 17
# Cada tabla tiene: fila de año, fila de encabezados, y luego meses con COMPRA/VENTA
COL_ANIOS = [2, 5, 8, 11, 14, 17]

for fila in range(33, ws.max_row + 1):
    # Detectar fila de año: col 1 vacía y col 2 tiene un número
    col1 = ws.cell(row=fila, column=1).value
    col2 = ws.cell(row=fila, column=2).value
    if col1 is not None or not isinstance(col2, (int, float)):
        continue

    # Fila de año encontrada: leer los años de cada bloque
    anios = {}
    for col_anio in COL_ANIOS:
        valor = ws.cell(row=fila, column=col_anio).value
        if isinstance(valor, (int, float)):
            anios[col_anio] = int(valor)

    # Leer las filas de meses (empiezan 3 filas después del año)
    fila_inicio_meses = fila + 3
    for fila_mes in range(fila_inicio_meses, fila_inicio_meses + 12):
        mes = ws.cell(row=fila_mes, column=1).value
        if mes not in MESES:
            continue
        for col_anio, anio in anios.items():
            compra = ws.cell(row=fila_mes, column=col_anio).value
            venta = ws.cell(row=fila_mes, column=col_anio + 1).value
            if compra is not None:
                datos.append([f"{anio} {mes}", compra, venta])

# Ordenar cronológicamente: separar encabezado del resto
encabezado = datos[0]
registros = datos[1:]

def clave_orden(registro):
    periodo = registro[0]
    if isinstance(periodo, (int, float)):
        # Dato anual (ej: 1945)
        return (int(periodo), 0)
    # Dato mensual (ej: "1970 Enero")
    partes = str(periodo).split(" ", 1)
    anio = int(partes[0])
    mes_idx = MESES.index(partes[1]) + 1 if partes[1] in MESES else 0
    return (anio, mes_idx)

registros.sort(key=clave_orden)

# Crear el archivo de resultado
wb_resultado = Workbook()
ws_resultado = wb_resultado.active
ws_resultado.title = "USD Fin Mes"

ws_resultado.append(encabezado)
for registro in registros:
    ws_resultado.append(registro)

# Guardar con nombre único usando timestamp
timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
nombre_archivo = f"clase21/resultados/cotizacion_dolar_{timestamp}.xlsx"
wb_resultado.save(nombre_archivo)

print(f"Archivo guardado: {nombre_archivo}")
print(f"Total de filas extraídas: {len(datos)} (1 encabezado + {len(datos) - 1} datos)")