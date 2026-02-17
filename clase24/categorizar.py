from openpyxl import Workbook, load_workbook
from pathlib import Path
import os

def categorizar_excel(path_archivo):
    wb = load_workbook(path_archivo)
    sheet = wb["Transacciones"]
    diccionarioCategoria = {}
    with open("clasificacion.txt","r",encoding="utf-8") as f:
        while True:
            linea = f.readline()
            if len(linea) == 0:
                break;
            partes = linea.split(",")
            descripcion = partes[0].strip()
            categoria = partes[1].strip()
            diccionarioCategoria[descripcion] = categoria
    
    lista_categorias = []
    with open("categorias.txt", "r",encoding="utf-8") as f: 
        while True:
            linea = f.readline().strip()
            if len(linea) == 0:
                break;
            lista_categorias.append(linea)
    with  open("clasificacion.txt", "a",encoding="utf-8") as archivo_clasificacion:
        with open("categorias.txt", "a",encoding="utf-8") as archivo_categorias:
            for fila in range(2, sheet.max_row):
                try:
                    fecha = sheet.cell(fila, 1).value
                    descripcion = sheet.cell(fila, 2).value
                    monto = sheet.cell(fila,3).value
                    tipo = sheet.cell(fila,4).value
                    if tipo == "Ingreso":
                        continue
                    if descripcion in diccionarioCategoria:
                        continue
                    os.system('cls')
                    print("para la descripcion:",descripcion,"\n¿Qué categoría corresponde?")
                    for indice, categoria in enumerate(lista_categorias):
                        print(categoria,"(",indice,")")
                    print("Ingresar nueva categoría","(",len(lista_categorias),")")            
                    categoria_elegida = int(input("escoge una categoría: "))
                    if categoria_elegida == len(lista_categorias):
                        os.system('cls')
                        texto = "ingresa el nombre de la categoría para" + descripcion + ": "
                        nombre_categoria = input(texto)
                        archivo_categorias.write("\n" +nombre_categoria)
                        lista_categorias.append(nombre_categoria)
                    diccionarioCategoria[descripcion] = lista_categorias[categoria_elegida]
                    linea = "\n" + descripcion + "," + diccionarioCategoria[descripcion] 
                    archivo_clasificacion.write(linea)
                except:
                    continue
            
        


path = Path("transacciones.xlsx")  
categorizar_excel(path)