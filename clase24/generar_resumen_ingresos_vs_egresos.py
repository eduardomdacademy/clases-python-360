from openpyxl import load_workbook
from pathlib import Path
import matplotlib.pyplot as plt
def generar_resumen_ingresos_vs_egresos(path_archivo):
    workbook = load_workbook(path_archivo)
    sheet = workbook["Transacciones"]
    total_ingresos = 0
    total_egresos = 0
    for fila in range(2,sheet.max_row):
        fecha = sheet.cell(fila, 1).value
        descripcion = sheet.cell(fila, 2).value
        monto = sheet.cell(fila,3).value
        tipo = sheet.cell(fila,4).value
        if tipo == "Ingreso":
            total_ingresos += monto
        elif tipo == "Gasto":
            total_egresos += monto
    print("total_ingresos", total_ingresos)
    print("total_egresos", total_egresos)

    categorias = ["Ingresos", "Egresos"]
    valores = [total_ingresos, total_egresos]

    plt.bar(categorias, valores, color=["green", "red"])
    plt.ticklabel_format(style="plain", axis="y")
    plt.title("Ingresos vs Egresos")
    plt.ylabel("Monto")
    plt.show()

path = Path("transacciones.xlsx")
generar_resumen_ingresos_vs_egresos(path)