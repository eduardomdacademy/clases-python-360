from categorizar import categorizar_excel
from openpyxl import load_workbook
from pathlib import Path
import matplotlib.pyplot as plt
def generar_resumen_por_categoria(path_archivo):
    categorizar_excel(path_archivo)
    totalesPorCategoria = {}
    with open("categorias.txt","r",encoding="utf-8") as archivo_categorias:
        while True:
            categoria = archivo_categorias.readline().strip()
            if len(categoria) == 0:
                break
            totalesPorCategoria[categoria] = 0
    diccionarioCategoria = {}
    with open("clasificacion.txt","r", encoding="utf-8") as archivo_clasificacion:
        while True:
            linea = archivo_clasificacion.readline()
            if len(linea) == 0:
                break
            partes = linea.split(",")
            descripcion = partes[0].strip()
            categoria = partes[1].strip()
            diccionarioCategoria[descripcion] = categoria
    workbook = load_workbook(path_archivo)
    sheet = workbook["Transacciones"]
    for fila in range(2, sheet.max_row):
        fecha = sheet.cell(fila, 1).value
        descripcion = sheet.cell(fila,2).value
        monto = sheet.cell(fila, 3).value
        tipo = sheet.cell(fila, 4).value
        if tipo == "Ingreso":
            continue
        categoria = diccionarioCategoria[descripcion]
        totalesPorCategoria[categoria] += monto

    categorias = list(totalesPorCategoria.keys())
    montos = list(totalesPorCategoria.values())

    total = sum(montos)
    plt.figure(figsize=(8, 8))
    plt.pie(montos, labels=categorias,
            autopct=lambda pct: f"Gs.{total * pct / 100:,.0f}\n({pct:.1f}%)")
    plt.title("Gastos por categoría")
    plt.show()
        
path = Path("transacciones.xlsx")
generar_resumen(path)